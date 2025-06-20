import streamlit as st
import pandas as pd
from io import BytesIO
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide")
st.title("📋 DTS Corrector")

# === Uploads ===
sf_file = st.file_uploader("Upload Start Form CSV (SFlist_*.csv)", type="csv")
ts_files = st.file_uploader("Upload Daily Time Sheet Excel files", type="xlsx", accept_multiple_files=True)

if sf_file and ts_files:
    if st.button("▶️ Process Uploaded Files"):
        with st.spinner("Processing..."):

            from datetime import datetime
            import re


            def clean_datum_column(df, file_name):
                today = datetime.today()

                # Extract fallback date from filename like "20250619"
                match = re.search(r"(20\d{6})", file_name)
                date_from_filename = None
                if match:
                    try:
                        date_from_filename = datetime.strptime(match.group(1), "%Y%m%d")
                    except:
                        pass

                def parse_or_fallback(value):
                    if pd.isna(value) or str(value).strip() == "":
                        return date_from_filename

                    str_val = str(value).strip()

                    # Format 1: YYYYMMDD (e.g. 20250619)
                    if re.fullmatch(r"20\d{6}", str_val):
                        try:
                            return datetime.strptime(str_val, "%Y%m%d")
                        except:
                            return date_from_filename

                    # Format 2: YYMMDD (e.g. 250619 → 2025-06-19)
                    if re.fullmatch(r"\d{6}", str_val):
                        try:
                            return datetime.strptime(str_val, "%y%m%d")
                        except:
                            return date_from_filename

                    # Format 3: try anything else (e.g. 19.06.2025)
                    try:
                        parsed = pd.to_datetime(str_val, dayfirst=True, errors="coerce")
                        return parsed if pd.notna(parsed) else date_from_filename
                    except:
                        return date_from_filename

                if "Dátum" in df.columns:
                    df["Dátum"] = df["Dátum"].apply(parse_or_fallback)
                else:
                    df["Dátum"] = date_from_filename

                return df


            # === Load SF list ===
            df_sf = pd.read_csv(sf_file)
            assert all(col in df_sf.columns for col in ["Crew list name", "Project job title", "Sf number"])
            sf_pairs = list(zip(df_sf["Crew list name"], df_sf["Project job title"], df_sf["Sf number"]))

            # === DTS Loader ===
            def load_dts_with_header_by_datum(file):
                preview = pd.read_excel(file, header=None, nrows=20, usecols="A")
                header_row = None
                for i, value in preview[0].items():
                    if str(value).strip().lower() == "dátum":
                        header_row = i
                        break
                if header_row is None:
                    raise ValueError(f"'Dátum' not found in first column of {file.name}")
                df = pd.read_excel(file, header=header_row, usecols="A:M")
                df.columns = df.columns.map(str)
                return df

            # === Load DTS files ===
            combined = []
            for file in ts_files:
                try:
                    df = load_dts_with_header_by_datum(file)
                    df["Source File"] = file.name
                    df = clean_datum_column(df, file.name)
                    combined.append(df)
                except Exception as e:
                    st.warning(f"⚠️ Could not load {file.name}: {e}")

            df_dts_all = pd.concat(combined, ignore_index=True)

            # === Drop rows where neither Dolgozott nor Kezdés indicate work
            def was_working(row):
                dolgozott = row.get("Dolgozott", None)
                if pd.isna(dolgozott) or str(dolgozott).strip() in ["", "0", "0.0"]:
                    worked = False
                else:
                    try:
                        worked = float(str(dolgozott).replace(",", ".").strip()) > 0
                    except ValueError:
                        worked = False

                kezdés = row.get("Kezdés", None)
                has_kezdes = pd.notna(kezdés) and str(kezdés).strip() != ""
                return worked or has_kezdes

            mask = df_dts_all.apply(was_working, axis=1)
            dropped_rows = df_dts_all[~mask].copy()
            df_dts = df_dts_all[mask].copy()


        mask = df_dts_all.apply(was_working, axis=1)
        dropped_rows = df_dts_all[~mask].copy()
        df_dts = df_dts_all[mask].copy()



        # === Matching functions ===
        def normalize(text): return str(text).strip().lower()

        def find_best_match(name, title):
            name_norm = normalize(name)
            title_norm = normalize(title)
            best_score, best_match = 0, None
            for sf_name, sf_title, sf_number in sf_pairs:
                name_score = fuzz.token_set_ratio(name_norm, normalize(sf_name))
                title_score = fuzz.token_sort_ratio(title_norm, normalize(sf_title))
                score = 0.7 * name_score + 0.3 * title_score
                if score > best_score:
                    best_score = score
                    best_match = (sf_name, sf_title, sf_number)
            return best_match if best_score >= 85 else None

        def match_blank_deal_title(title):
            df_bd = df_sf[df_sf["Sf number"].astype(str).str.startswith("BD")]
            best_score, best_title, best_sf_number = 0, None, None
            for _, row in df_bd.iterrows():
                score = fuzz.token_sort_ratio(normalize(title), normalize(row["Project job title"]))
                if score > best_score:
                    best_score = score
                    best_title = row["Project job title"]
                    best_sf_number = row["Sf number"]
            return (best_title, best_sf_number) if best_score >= 85 else (None, None)


        from datetime import time


        def parse_excel_time(val):
            if pd.isna(val):
                return None
            str_val = str(val).strip().replace(",", ".").replace(" ", "")

            # Try parsing time in different formats
            for fmt in ("%H:%M", "%H.%M", "%H:%M:%S"):
                try:
                    parsed = datetime.strptime(str_val, fmt)
                    return parsed.time()
                except:
                    continue
            return None


        # === Match and Correct ===
        results = {
            "Original Name": [], "Original Title": [],
            "Név (angolul)": [], "Beosztás": [],
            "Match result": [], "Matched SF number": []
        }

        for _, row in df_dts.iterrows():
            orig_name, orig_title = row["Név (angolul)"], row["Beosztás"]
            results["Original Name"].append(orig_name)
            results["Original Title"].append(orig_title)

            match = find_best_match(orig_name, orig_title)
            if match:
                m_name, m_title, m_sf = match
                results["Név (angolul)"].append(m_name)
                results["Beosztás"].append(m_title)
                results["Matched SF number"].append(m_sf)
                results["Match result"].append("Changed" if (m_name != orig_name or m_title != orig_title) else "Same")
            else:
                bd_title, bd_sf = match_blank_deal_title(orig_title)
                if bd_title:
                    results["Név (angolul)"].append(orig_name)
                    results["Beosztás"].append(bd_title)
                    results["Matched SF number"].append(bd_sf)
                    results["Match result"].append("Blank Deal")
                else:
                    results["Név (angolul)"].append(orig_name)
                    results["Beosztás"].append(orig_title)
                    results["Matched SF number"].append("")
                    results["Match result"].append("Same")

        # === Output Excel File in Memory ===
        df_result = df_dts.copy()
        # --- Define time-based columns that should be properly formatted for Excel
        time_columns = ["Kezdés", "Végzés", "Utolsó Végzés"]



        for col, values in results.items():
            df_result[col] = values

        # Convert time columns to Excel-friendly time objects
        for col in time_columns:
            if col in df_result.columns:
                df_result[col] = df_result[col].apply(parse_excel_time)



        # Original column names (from your data)
        ot_chart_columns = [
            "Dátum", "Név (angolul)", "Beosztás", "Beosztás megjegyzés",
            "Dolgozott", "Kezdés", "Végzés", "Megjegyzés", "Helyszín",
            "Unit", "Ebédidő megváltás", "Utolsó Végzés", "Munkanap", "Source File"
        ]

        # Corresponding renamed columns for OT Chart
        ot_chart_renames = {
            "Dátum": "Date",
            "Név (angolul)": "Name",
            "Beosztás": "Title",
            "Beosztás megjegyzés": "Title Mod",
            "Dolgozott": "Wx",
            "Kezdés": "Start",
            "Végzés": "End",
            "Megjegyzés": "Note",
            "Helyszín": "Location",
            "Unit": "Unit",
            "Ebédidő megváltás": "F_MP",
            "Utolsó Végzés": "LW_O",
            "Munkanap": "WD_O",
            "Source File": "File"
        }

        df_ot_chart = df_result[[col for col in ot_chart_columns if col in df_result.columns]].copy()
        df_ot_chart.rename(columns=ot_chart_renames, inplace=True)

        for col in time_columns:
            if col in df_ot_chart.columns:
                df_ot_chart[col] = df_ot_chart[col].apply(parse_excel_time)

        from datetime import datetime, time as dtime


        def time_to_float(t):
            if isinstance(t, time):
                return (t.hour * 60 + t.minute) / (24 * 60)
            return t  # Leave as-is if already float or None


        for col in ["Start", "End", "LW_O"]:
            if col in df_ot_chart.columns:
                df_ot_chart[col] = df_ot_chart[col].apply(time_to_float)

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_ot_chart.to_excel(writer, index=False, sheet_name="For OT Chart")
            df_result.to_excel(writer, index=False, sheet_name="Corrected TS")
            dropped_rows.to_excel(writer, index=False, sheet_name="Dropped Rows")

            wb = writer.book

            from uuid import uuid4

            for sheet_name, df_sheet in [("For OT Chart", df_ot_chart), ("Corrected TS", df_result),
                                         ("Dropped Rows", dropped_rows)]:
                ws = wb[sheet_name]

                # === Only format "For OT Chart"

                if sheet_name == "For OT Chart":
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        for cell in row:
                            if cell.column_letter == "A":  # Only format date
                                cell.number_format = "yyyy.mm.dd"
                            # Leave other cells untouched (let them stay as floats)

                # Add Excel table
                if not df_sheet.empty and df_sheet.shape[1] > 0 and ws.max_row > 1:
                    max_row = ws.max_row
                    max_col = ws.max_column
                    last_col = get_column_letter(max_col)
                    unique_name = f"Table_{sheet_name.replace(' ', '_')}_{uuid4().hex[:4]}"
                    table = Table(displayName=unique_name, ref=f"A1:{last_col}{max_row}")
                    table.tableStyleInfo = TableStyleInfo(
                        name="TableStyleLight1",
                        showFirstColumn=False, showLastColumn=False,
                        showRowStripes=False, showColumnStripes=False
                    )
                    ws.add_table(table)

                # Auto column width
                for col in ws.columns:
                    max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        st.success("✅ Correction complete!")
        st.download_button("📥 Download Corrected Excel",
                           data=output.getvalue(),
                           file_name="Corrected_Combined_Daily_Timesheets.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

